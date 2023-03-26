import os
import json
import pandas as pd
import httpx
import asyncio
from itertools import chain
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

print("timer started now!")
start_time = time.perf_counter()

#enable 1.21 gigawatts...
async def get_ip_data(ip_address):
    async with httpx.AsyncClient() as client:
        headers = {'Token': '<token here>'}
        response = await client.get(f"https://api.spur.us/v2/context/{ip_address}", headers=headers)
        print(response.json())
        if response.status_code == 200:
            return response.json()
        else:
            return None

def process_ip_data(ip_data):
    if ip_data is None:
        return {}

    flattened_data = {}

    #flatten json (cater for strings, dicts and lists)
    def flatten_json(data, prefix=''):
        flattened_data_inner = {}

        if isinstance(data, str):
            return flattened_data_inner

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, dict) or isinstance(value, list):
                    flattened_data_inner.update(flatten_json(value, f"{prefix}{key}."))
                else:
                    flattened_data_inner[f"{prefix}{key}"] = value
        elif isinstance(data, list):
            for i, item in enumerate(data):
                if isinstance(item, dict) or isinstance(item, list):
                    flattened_data_inner.update(flatten_json(item, f"{prefix}{i}."))
                else:
                    flattened_data_inner[f"{prefix}{i}"] = item

        return flattened_data_inner

    flattened_data.update(flatten_json(ip_data))
    return flattened_data

async def main():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    file_path = os.path.join(desktop_path, "ips.txt")

    with open(file_path, "r") as file:
        ip_addresses = [ip.strip() for ip in file.readlines()]

    # Fetch IP data asynchronously
    ip_data = await asyncio.gather(*[get_ip_data(ip) for ip in ip_addresses])

    # Process IP data
    processed_data = [process_ip_data(data) for data in ip_data]

    # Create a dataframe from the processed data
    df = pd.DataFrame(processed_data)

    # Export the dataframe to an xlsx file
    output_file_path = os.path.join(desktop_path, "ip_data.xlsx")
    
    #df.to_excel(output_file_path, index=False, engine="openpyxl")

        # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Set the header row
    for col, header in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write the data to the worksheet and apply conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    vpn_keyword = "VPN"

    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        vpn_detected = any(vpn_keyword in str(cell) for cell in row)
        for col_index, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            if vpn_detected: cell.fill = red_fill
            # Adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook
    wb.save(output_file_path)

asyncio.run(main())

#finish timer
end_time = time.perf_counter()
elapsed_time = end_time - start_time
print(f"The script took {elapsed_time:.4f} seconds to execute.")